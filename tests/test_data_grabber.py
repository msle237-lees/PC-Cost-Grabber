"""
@file test_data_grabber.py
@brief Unit tests for data_grabber module (parallel fetch → JSON/Excel/HTTP).

@details
  - Mocks pcpartpicker.API and requests.Session for deterministic tests.
  - Swaps ProcessPoolExecutor with a dummy in-process executor to avoid spawning.
  - Validates:
      * _worker_fetch() → returns JSON-ready rows
      * fetch_all() → aggregates all supported types
      * write_json() → writes a JSON file with expected structure
      * write_excel() → creates an .xlsx with a sheet per type (skips if pandas/openpyxl missing)
      * post_results() → posts each item, aggregates success/failure

@requirements
  - pytest
  - pandas, openpyxl (only if you want the Excel test to run; otherwise it will be skipped)

@note
  These tests avoid network and multiprocessing to remain fast and reliable.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from types import ModuleType
from concurrent.futures import Future  # ✅ use a real Future so as_completed() works

# Add the parent directory to sys.path so we can import data_grabber
sys.path.insert(0, str(Path(__file__).parent.parent))

from typing import Any, Dict, List  # isort: skip
from datetime import datetime  # isort: skip
from decimal import Decimal  # isort: skip

import pytest


# -----------------------------
# Helpers / Fakes
# -----------------------------

class _FakePart:
    """A non-JSON-serializable object to ensure _to_jsonable() is used."""
    def __init__(self, name: str, price: Decimal, when: datetime, meta: Dict[str, Any] | None = None):
        self.name = name
        self.price = price
        self.when = when
        self.meta = meta or {"nested": {"ok": True}}


class _FakeAPI:
    """Fake pcpartpicker.API that returns dict keyed by component type."""
    def __init__(self) -> None:
        self.calls: List[str] = []

    def retrieve(self, component_type: str) -> Dict[str, List[Any]]:
        self.calls.append(component_type)
        # Each type returns two fake parts
        return {
            component_type: [
                _FakePart(f"{component_type}-A", Decimal("99.99"), datetime(2024, 1, 1, 12, 0, 0)),
                _FakePart(f"{component_type}-B", Decimal("149.50"), datetime(2024, 2, 2, 13, 30, 0)),
            ]
        }


class _DummyExecutor:
    """
    In-process stand-in for ProcessPoolExecutor.

    Runs the submitted callable immediately in the current process and returns a
    real concurrent.futures.Future with its result set. This keeps behavior
    compatible with concurrent.futures.as_completed().
    """
    def __init__(self, max_workers=None):
        self.max_workers = max_workers
        self._futures: List[Future] = []

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        return False

    def submit(self, fn, *args, **kwargs) -> Future:
        fut: Future = Future()
        try:
            result = fn(*args, **kwargs)
            fut.set_result(result)
        except Exception as e:  # noqa: BLE001
            fut.set_exception(e)
        self._futures.append(fut)
        return fut


def _install_fake_pcpartpicker(monkeypatch: pytest.MonkeyPatch) -> None:
    """
    Install a fake 'pcpartpicker' module into sys.modules so that
    'from pcpartpicker import API' works inside data_grabber._worker_fetch.
    """
    fake_mod = ModuleType("pcpartpicker")
    fake_mod.API = _FakeAPI
    monkeypatch.setitem(sys.modules, "pcpartpicker", fake_mod)


# -----------------------------
# Fixtures
# -----------------------------

@pytest.fixture(autouse=True)
def _patch_executor_and_module(monkeypatch: pytest.MonkeyPatch):
    """
    Auto-applied fixture to:
    - Replace data_grabber.ProcessPoolExecutor with an in-process dummy.
    - Ensure a fake pcpartpicker module is present.
    """
    import data_grabber  # the module under test

    _install_fake_pcpartpicker(monkeypatch)
    monkeypatch.setattr(data_grabber, "ProcessPoolExecutor", _DummyExecutor)
    yield


# -----------------------------
# Tests
# -----------------------------

def test_worker_fetch_returns_jsonable_rows(monkeypatch: pytest.MonkeyPatch):
    """
    @test Ensure _worker_fetch returns (type, list[dict]) with JSON-safe values.
    """
    import data_grabber as dg

    comp_type = "cpu"
    k, rows = dg._worker_fetch(comp_type)
    assert k == comp_type
    assert isinstance(rows, list)
    assert len(rows) == 2

    first = rows[0]
    # Ensure fields we expect were converted to JSON-safe strings
    assert first["name"] == "cpu-A"
    assert isinstance(first["price"], str) and first["price"] == "99.99"
    assert isinstance(first["when"], str) and first["when"].startswith("2024-01-01T12:00:00")
    assert first["meta"]["nested"]["ok"] is True


def test_fetch_all_aggregates_all_supported_types(monkeypatch: pytest.MonkeyPatch):
    """
    @test fetch_all submits a job per supported type and builds the mapping.
    """
    import data_grabber as dg

    results = dg.fetch_all(max_workers=2)
    # Every supported type should be present
    for t in dg.supported_types:
        assert t in results
        assert isinstance(results[t], list)
        assert len(results[t]) == 2


def test_write_json_writes_file(tmp_path):
    """
    @test write_json writes pretty JSON to disk with safe values.
    """
    import data_grabber as dg

    # Small subset to keep the file short
    sample = {
        "cpu": [{"name": "cpu-1", "price": "10.00", "when": "2024-01-01T00:00:00"}],
        "gpu": [{"name": "gpu-1", "price": "20.00", "when": "2024-02-02T00:00:00"}],
    }
    out = tmp_path / "out.json"
    dg.write_json(sample, str(out))

    assert out.exists()
    data = json.loads(out.read_text(encoding="utf-8"))
    assert set(data.keys()) == {"cpu", "gpu"}
    assert data["cpu"][0]["price"] == "10.00"


@pytest.mark.skipif(
    pytest.importorskip("pandas") is None or pytest.importorskip("openpyxl") is None,
    reason="pandas/openpyxl required for Excel test",
)
def test_write_excel_creates_sheets(tmp_path):
    """
    @test write_excel creates an .xlsx with a sheet per type.
    """
    import pandas as pd  # noqa: F401
    import data_grabber as dg

    # Minimal realistic results
    sample = {
        "cpu": [{"name": "cpu-1", "price": "10.00"}],
        "video-card": [{"name": "gpu-1", "price": "20.00"}],
        "speakers": [],  # empty still creates a sheet
    }
    xlsx = tmp_path / "parts.xlsx"
    dg.write_excel(sample, str(xlsx))
    assert xlsx.exists()

    # Optional: open and verify sheet names/contents
    import openpyxl  # type: ignore
    wb = openpyxl.load_workbook(str(xlsx))
    # Sheet names sanitized and limited
    assert "cpu" in wb.sheetnames
    assert "video-card" in wb.sheetnames
    assert "speakers" in wb.sheetnames

    # Check a cell value
    ws = wb["cpu"]
    # Header row starts at A1, data at A2
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert "name" in headers
    assert "price" in headers


def test_post_results_sends_requests_and_summarizes(monkeypatch: pytest.MonkeyPatch):
    """
    @test post_results posts each item and aggregates success/failure counts.
    """
    import data_grabber as dg

    # Build fake results: 3 items total
    results = {
        "cpu": [{"id": 1}, {"id": 2}],
        "mouse": [{"id": 3}],
        "speakers": {"error": "fetch failed"},  # should be skipped
    }

    # Fake requests.Session
    class _Resp:
        def __init__(self, code: int, text: str = "ok"):  # noqa: D401
            self.status_code = code
            self.text = text

    class _FakeSession:
        def __init__(self):
            self.posts: List[dict] = []

        def post(self, url, json=None, timeout=10.0, headers=None):
            self.posts.append({"url": url, "json": json})
            # Make one of them fail to test error handling
            if json and json.get("id") == 2:
                return _Resp(500, "boom")
            return _Resp(201, "created")

    fake_session = _FakeSession()

    def _fake_session_ctor():
        return fake_session

    # Patch requests.Session to our fake
    fake_requests = ModuleType("requests")
    fake_requests.Session = _fake_session_ctor  # type: ignore[attr-defined]
    monkeypatch.setitem(sys.modules, "requests", fake_requests)

    summary = dg.post_results(
        results=results,
        base_url="http://localhost:8000",
        endpoint_template="/pcparts/{type}",
        concurrency=4,
        timeout=5.0,
        headers={"X-Test": "1"},
    )

    # We had 3 posts; one failed (id=2)
    assert summary["posted"] == 2
    assert summary["failed"] == 1
    assert len(summary["errors"]) == 1

    # Verify URLs formed with {type} placeholder
    urls = [p["url"] for p in fake_session.posts]
    assert "http://localhost:8000/pcparts/cpu" in urls
    assert "http://localhost:8000/pcparts/mouse" in urls
